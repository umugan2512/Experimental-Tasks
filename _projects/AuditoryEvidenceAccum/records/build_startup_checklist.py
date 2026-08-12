# !/usr/bin/python3
# -*- coding: utf-8 -*-
"""
Generates session_startup_checklist.xlsx and session_startup_checklist.md -- a single-page,
printable/duplicable form (not a data table) for physical record-keeping: whoever runs a session
fills in animal ID, which protocol was run, handwritten weight/weight%/reward-consumed, ticks off
an equipment-on/off checklist and a set of pre-session checks, then fills in an outcome summary at
the end. Meant to be printed (or duplicated as a new sheet/copy per session) the same way a
physical lab notebook page would be -- purely a blank template, never auto-populated (there's no
digital source for any of this; contrast with build_training_log.py, which reads real session
logs).

Both formats are generated from the single CHECKLIST_SECTIONS structure below so the checklist
text only ever needs editing in one place. The .md version uses GitHub-flavored checkbox syntax
(`- [ ]`) -- renders as real checkboxes in a Markdown preview (VS Code, Typora, GitHub, a browser
extension, ...) which is the easiest path to a clean print; opening the raw .md in a plain text
editor and printing it still shows readable "[ ]" boxes.

Run once (or re-run any time you want to reset either file back to blank):
    /c/Users/2P-Behav/.conda/envs/pybpod-environment/python.exe build_startup_checklist.py
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_RECORDS_DIR = os.path.dirname(os.path.abspath(__file__))
_XLSX_OUTPUT_PATH = os.path.join(_RECORDS_DIR, 'session_startup_checklist.xlsx')
_MD_OUTPUT_PATH = os.path.join(_RECORDS_DIR, 'session_startup_checklist.md')

TITLE = 'Session Startup Checklist -- AuditoryEvidenceAccum'

# Single source of truth for both output formats. Each section is (title, [(kind, text), ...])
# with kind in {'blank', 'blank_multiline', 'checkbox'}.
CHECKLIST_SECTIONS = [
    ('Session identification', [
        ('blank', 'Date:'),
        ('blank', 'Animal ID:'),
        ('blank', 'Experimenter:'),
        ('blank', 'Protocol run (e.g. "Stage 2 threshold staircase"):'),
        ('blank', 'Time started:'),
        ('blank', 'Time ended:'),
    ]),
    ('Weight and reward (handwritten)', [
        ('blank', 'Weight (g):'),
        ('blank', 'Weight % of baseline:'),
        ('blank', 'Reward consumed (count or est. volume):'),
    ]),
    ('Equipment on (before starting)', [
        ('checkbox', 'Bpod board powered on and connected'),
        ('checkbox', 'Rotary encoder module connected'),
        ('checkbox', 'HiFi module connected (if this protocol uses sound)'),
        ('checkbox', 'Monitor / dot display connected and positioned'),
        ('checkbox', 'Water line primed and checked for drips'),
        ('checkbox', 'Spout position confirmed against the training-stage schedule'),
    ]),
    ('Pre-session checks', [
        ('checkbox', 'Animal health check normal (grooming, posture, behavior)'),
        ('checkbox', 'Water restriction on schedule (checked against the training-log workbook)'),
        ('checkbox', 'Wheel spins freely, no obstruction'),
        ('checkbox', 'Live plot window opened and confirmed showing data'),
    ]),
    ('Equipment off (after finishing)', [
        ('checkbox', 'Bpod session stopped cleanly (not Killed)'),
        ('checkbox', 'HiFi module powered off (if used)'),
        ('checkbox', 'Monitor / dot display powered off'),
        ('checkbox', 'Water line checked, no residual dripping'),
        ('checkbox', 'Animal returned to home cage, weighed if required'),
    ]),
    ('Outcome summary (fill in after the session)', [
        ('blank', 'Trial count:'),
        ('blank', 'Advance-ready (Y/N, from console output):'),
        ('blank', 'Any errors / crashes / hardware issues:'),
        ('blank_multiline', 'Notes:'),
        ('blank', 'Signature / initials:'),
    ]),
]


# --- xlsx ------------------------------------------------------------------------------------

_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11)
_LABEL_FONT = Font(bold=True)
_THIN = Side(style='thin', color='999999')
_BOX_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
_XLSX_CHECKBOX = '☐'   # hand-marked with an X or a checkmark when printed/filled in


def _xlsx_section(ws, row, title):
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = _SECTION_FILL
    return row + 1


def _xlsx_label_blank(ws, row, label, blank_width=3):
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    for c in range(2, 2 + blank_width):
        ws.cell(row=row, column=c).border = _BOX_BORDER
    return row + 1


def _xlsx_checkbox_line(ws, row, text):
    ws.cell(row=row, column=1, value=_XLSX_CHECKBOX)
    ws.cell(row=row, column=2, value=text)
    return row + 1


def build_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Startup checklist'

    row = 1
    ws.cell(row=row, column=1, value=TITLE).font = _TITLE_FONT
    row += 2

    for title, items in CHECKLIST_SECTIONS:
        row = _xlsx_section(ws, row, title)
        for kind, text in items:
            if kind == 'checkbox':
                row = _xlsx_checkbox_line(ws, row, text)
            elif kind == 'blank_multiline':
                row = _xlsx_label_blank(ws, row, text, blank_width=4)
                for extra in range(3):
                    for c in range(2, 6):
                        ws.cell(row=row + extra, column=c).border = _BOX_BORDER
                row += 3
            else:   # 'blank'
                blank_width = 4 if len(text) > 30 else 3
                row = _xlsx_label_blank(ws, row, text, blank_width=blank_width)
        row += 1

    ws.column_dimensions['A'].width = 42
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 16
    ws.sheet_view.showGridLines = False
    ws.print_area = 'A1:F{0}'.format(row + 1)

    wb.save(_XLSX_OUTPUT_PATH)
    print("Wrote {0}".format(_XLSX_OUTPUT_PATH), flush=True)


# --- markdown ----------------------------------------------------------------------------------

_MD_BLANK_LINE = '_' * 32


def build_md():
    lines = ['# {0}'.format(TITLE), '']
    for title, items in CHECKLIST_SECTIONS:
        lines.append('## {0}'.format(title))
        lines.append('')
        for kind, text in items:
            if kind == 'checkbox':
                lines.append('- [ ] {0}'.format(text))
            elif kind == 'blank_multiline':
                lines.append('**{0}**'.format(text))
                lines.append('')
                for _ in range(4):
                    lines.append(_MD_BLANK_LINE)
                    lines.append('')
            else:   # 'blank'
                lines.append('**{0}** {1}'.format(text, _MD_BLANK_LINE))
                lines.append('')   # blank line so consecutive fields don't collapse into one
                                    # paragraph under strict CommonMark rendering
        lines.append('')

    with open(_MD_OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print("Wrote {0}".format(_MD_OUTPUT_PATH), flush=True)


if __name__ == '__main__':
    build_xlsx()
    build_md()
