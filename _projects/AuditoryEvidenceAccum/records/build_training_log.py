# !/usr/bin/python3
# -*- coding: utf-8 -*-
"""
Scans every session CSV under experiments/*/setups/*/sessions/ **on this machine** and
update-in-place merges the results into training_log.xlsx -- one sheet per animal, sessions
grouped into contiguous per-stage blocks (Stage 1's rows together, then Stage 2's, ...) below a
hand-editable DOB/Sex/Strain/Baseline-weight header block.

**Session merging**: a session that got Stopped, Killed, or crashed partway through and was
restarted within an hour (same protocol) is combined into ONE row + one combined data-output file
-- see `group_sessions()`. A session that reached its own natural end (`VAR_MAX_TRIALS`) is never
merged with what follows. Applies to every subject uniformly, including bench-test ones.

**Update-in-place, not rebuild-from-scratch** -- this file is meant to be shared across multiple
behavior-box computers via git (each box sees only its own local session data for its own
animal(s)). A run on this machine:
  - loads the existing training_log.xlsx if present (e.g. just `git pull`-ed),
  - touches ONLY the sheet(s) for subjects it found sessions for locally,
  - leaves every other subject's sheet in the workbook byte-for-byte untouched (safe for Box A to
    run this without ever seeing, let alone clobbering, Box B's/C's animals),
  - regenerates its own subject's data rows fresh every run from (current local scan + whatever
    Weight (g)/Notes values were already hand-typed into the sheet) -- necessary because merging
    can change which rows exist between runs (a restart appearing later can absorb a session that
    used to stand alone), and because rows are grouped into per-stage sections rather than a flat
    append-only list, so a newly-appearing stage's rows may need to land earlier in the sheet than
    rows already there. See `_rewrite_subject_sheet()`.
  - NEVER touches the `Weight (g)`/`Notes` values of a row that already had them, or the
    DOB/Sex/Strain/Baseline-weight header cells once they've been hand-filled -- those are the
    genuinely manual, hand-typed-in-Excel fields.

Parses the native Bpod STATE/EVENT/VAL structure directly out of the raw CSV rows (see
CLAUDE.md's "native-STATE/EVENT-parsing" note) via the shared `session_csv_parser` module (also
used by `session_struct_export.py`).

Workflow per box, per day: `git pull` -> run this script -> hand-fill any new blank `Weight (g)`
cells for today's sessions -> `git add training_log.xlsx && git commit && git push`.

Run with the pybpod-environment interpreter:
    /c/Users/2P-Behav/.conda/envs/pybpod-environment/python.exe build_training_log.py
"""
import glob
import math
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_RECORDS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_DIR = os.path.abspath(os.path.join(_RECORDS_DIR, '..'))
_OUTPUT_PATH = os.path.join(_RECORDS_DIR, 'training_log.xlsx')

sys.path.insert(0, os.path.join(_PROJECT_DIR, 'tasks', '_wheel_shaping_shared'))
import session_csv_parser     # noqa: E402 -- same-project import (records/ -> tasks/_wheel_shaping_shared/)
import session_struct_export  # noqa: E402
import staircase               # noqa: E402

_SIMPLE_GATES_PROTOCOLS = ('stage2_threshold_staircase',)   # which protocols have a
                                                              # stage2_simple_gates_met()-style
                                                              # advancement check worth computing

_MERGE_GAP = 3600.0   # seconds -- restarts within this long of a non-'completed' session's own end
                       # are combined into one row/struct with it (same protocol only)


def summarize_session(path):
    """ Parses one session CSV and returns a flat dict of everything the training log wants for
    one row, or None if the file has no recognizable protocol name OR the protocol has no
    PROTOCOL_CONFIG entry. The latter matters now that scan_all_sessions() also scans the Tests
    project (see its own docstring) -- that folder has many bench-test protocols (dot_wheel_test,
    camera_test, lick_reward, ...) with no meaningful outcome taxonomy to report; only protocols
    explicitly configured (currently: stage1/2 and full_protocol_lookback_test) generate a row, so
    the log doesn't fill up with all-'unknown'-outcome noise from every bench script ever run. """
    info, session_vals, trials = session_csv_parser.parse_session_csv(path)
    protocol_name = info.get('PROTOCOL-NAME')
    if not protocol_name or protocol_name not in session_csv_parser.PROTOCOL_CONFIG:
        return None
    config = session_csv_parser.PROTOCOL_CONFIG[protocol_name]

    subject = session_csv_parser.parse_subject_name(info.get('SUBJECT-NAME', ''))
    session_started = info.get('SESSION-STARTED', '')
    date = session_started.split(' ')[0] if session_started else ''

    trials = session_csv_parser.real_trials(trials)

    reward_count = withheld_count = no_movement_count = 0
    incorrect_count = wheel_abort_count = 0
    l_count = r_count = lick_count = consumed_reward_count = 0

    for trial in trials:
        outcome, side, _reward_duration, consumed = session_csv_parser.classify_trial(trial, config)
        if outcome == 'rewarded':
            reward_count += 1
            if consumed:
                consumed_reward_count += 1
        elif outcome == 'withheld':
            withheld_count += 1
        elif outcome == 'no_movement':
            no_movement_count += 1
        elif outcome == 'incorrect':
            incorrect_count += 1
        elif outcome == 'aborted':
            wheel_abort_count += 1
        if side == 'L':
            l_count += 1
        elif side == 'R':
            r_count += 1
        lick_count += len(trial['events'].get('Port1In', []))

    all_event_times = [t for trial in trials for times in trial['events'].values() for t in times]
    session_duration_s = (max(all_event_times) - min(all_event_times)) if all_event_times else None

    threshold_start_deg = session_csv_parser.find_val_forward(trials, session_vals, 'THRESHOLD_DEG')
    threshold_end_deg = session_csv_parser.find_val_backward(trials, session_vals, 'THRESHOLD_DEG')
    threshold_final_deg = 35.0   # matches every stage script's VAR_THRESHOLD_FINAL_DEG
    gain_mult_end = session_csv_parser.find_val_backward(trials, session_vals, 'GAIN_MULT')
    direction_ratio_end = session_csv_parser.find_val_backward(trials, session_vals, 'DIRECTION_RATIO')
    reward_ul = session_csv_parser.find_val_backward(trials, session_vals, 'REWARD_UL')
    aborts = session_csv_parser.sum_val(trials, 'QUIESCENCE_BREAKS')
    session_end_reason = session_csv_parser.find_val_str_backward(trials, session_vals,
                                                                    'SESSION_END_REASON')

    # ITI is a native STATE (the 'ITI' state's own logged duration), not a custom VAL row --
    # constant within a session in both stages, so any visited trial's value works; take the last.
    iti_end_s = None
    for trial in reversed(trials):
        if 'ITI' in trial['states']:
            start = trial['states']['ITI'][0]
            if not (start is None or math.isnan(start)):
                iti_end_s = trial['states']['ITI'][2]
                break

    simple_gates_met = None
    if protocol_name in _SIMPLE_GATES_PROTOCOLS and iti_end_s is not None:
        simple_gates_met = staircase.stage2_simple_gates_met(
            trial_count=len(trials), iti_s=iti_end_s,
            direction_ratio_in_band=(direction_ratio_end is not None and
                                      0.30 <= direction_ratio_end <= 0.70))

    started_dt = session_csv_parser.parse_datetime(session_started)
    ended_dt = session_csv_parser.session_end_datetime(info, started_dt, session_duration_s)
    time_of_day = started_dt.strftime('%H:%M') if started_dt is not None else ''

    consumed_volume_ul = (reward_ul * consumed_reward_count) if reward_ul is not None else None

    struct_mat_path = os.path.splitext(path)[0] + '_struct.mat'
    session_struct_path = struct_mat_path if os.path.exists(struct_mat_path) else None

    return {
        'session_started': session_started,
        'started_dt': started_dt,
        'ended_dt': ended_dt,
        'session_end_reason': session_end_reason,
        'subject': subject,
        'date': date,
        'time_of_day': time_of_day,
        'protocol': protocol_name,
        'trial_count': len(trials),
        'session_duration_s': session_duration_s,
        'reward_count': reward_count,
        'consumed_reward_count': consumed_reward_count,
        'withheld_count': withheld_count,
        'no_movement_count': no_movement_count,
        'incorrect_count': incorrect_count,
        'wheel_abort_count': wheel_abort_count,
        'l_count': l_count,
        'r_count': r_count,
        'lick_count': lick_count,
        'aborts': aborts,
        'reward_ul': reward_ul,
        'consumed_volume_ul': consumed_volume_ul,
        'threshold_start_deg': threshold_start_deg,
        'threshold_end_deg': threshold_end_deg,
        'threshold_final_deg': threshold_final_deg,
        'iti_end_s': iti_end_s,
        'gain_mult_end': gain_mult_end,
        'direction_ratio_end': direction_ratio_end,
        'simple_gates_met': simple_gates_met,
        'session_csv_path': path,
        'session_struct_path': session_struct_path,
    }


# --- scan every session on THIS machine ----------------------------------------------------------

_SCAN_PROJECT_DIRS = [
    _PROJECT_DIR,                                          # AuditoryEvidenceAccum -- real training
    os.path.join(_PROJECT_DIR, '..', 'Tests'),              # bench-test sessions -- see
                                                             # summarize_session()'s own docstring
                                                             # for why only EXPLICITLY configured
                                                             # protocols from here generate a row
]


def scan_all_sessions():
    """ Returns {subject: [session_summary, ...]}, each subject's list sorted by session_started.
    Only scans this machine's own local experiments/ folders (AuditoryEvidenceAccum + Tests, see
    _SCAN_PROJECT_DIRS) -- inherently per-box. """
    by_subject = {}
    for project_dir in _SCAN_PROJECT_DIRS:
        pattern = os.path.join(project_dir, 'experiments', '*', 'setups', '*', 'sessions', '*',
                                '*.csv')
        for path in sorted(glob.glob(pattern)):
            try:
                summary = summarize_session(path)
            except Exception as err:
                print("WARNING: failed to parse {0}: {1}".format(path, err), flush=True)
                continue
            if summary is None:
                continue
            by_subject.setdefault(summary['subject'], []).append(summary)
    for sessions in by_subject.values():
        sessions.sort(key=lambda s: s['session_started'])
    return by_subject


# --- session merging (restart-combining) ---------------------------------------------------------

def group_sessions(sessions):
    """ Returns a list of groups (each a chronological list of >=1 session summaries). A session
    whose `session_end_reason` is 'completed' (reached VAR_MAX_TRIALS naturally) always ends a
    group -- it's never combined with whatever runs next. Any other ending (Stop, Kill, or a
    crash -- `session_end_reason` absent) is eligible to merge with the next same-protocol session
    if it starts within `_MERGE_GAP` seconds of this one's own end. Applies to every subject
    uniformly, including bench-test ones -- merging reflects how the sessions actually happened,
    regardless of which GUI subject they were run under. """
    groups = []
    for session in sessions:
        if groups:
            last = groups[-1][-1]
            mergeable = (
                last['protocol'] == session['protocol']
                and last['session_end_reason'] != 'completed'
                and last['ended_dt'] is not None and session['started_dt'] is not None
                and (session['started_dt'] - last['ended_dt']).total_seconds() <= _MERGE_GAP
            )
            if mergeable:
                groups[-1].append(session)
                continue
        groups.append([session])
    return groups


def combine_group(group):
    """ Combines one group (>=1 raw session summaries) into a single row dict. A singleton group
    just carries its one session's values through unchanged; a real merge sums the per-trial
    tallies, including each member's own event-span `session_duration_s` (NOT the wall-clock gap
    between them -- summing active durations means a restart doesn't silently double-count idle
    time, and keeps a singleton group's Duration identical to that one session's own value, no
    behavior change for the common unmerged case), and takes threshold_start from the first
    member / everything else "as of end of session" from the last member (same semantics as a
    single session's own "_end" fields, just now "end of bout"). `session_struct_path` is left
    unset here -- the caller fills it in (single session: that session's own struct path;
    multi-session: the path `session_struct_export.merge_session_structs()` returns, since
    building that file needs filesystem access this pure function deliberately doesn't do). """
    first, last = group[0], group[-1]

    durations = [s['session_duration_s'] for s in group if s['session_duration_s'] is not None]
    duration_s = sum(durations) if durations else None

    consumed_vals = [s['consumed_volume_ul'] for s in group if s['consumed_volume_ul'] is not None]

    return {
        'session_started': '; '.join(s['session_started'] for s in group),
        'protocol': first['protocol'],
        'date': first['date'],
        'time_of_day': first['time_of_day'],
        # blank (not 1) for an ordinary, un-merged row -- "how many sessions were combined" only
        # means something once an actual merge happened; a bare "1" there would read as if
        # something got combined when nothing did.
        'num_sessions': len(group) if len(group) > 1 else None,
        'trial_count': sum(s['trial_count'] for s in group),
        'session_duration_s': duration_s,
        'reward_ul': last['reward_ul'],
        'consumed_volume_ul': sum(consumed_vals) if consumed_vals else None,
        'aborts': sum(s['aborts'] for s in group),
        'withheld_count': sum(s['withheld_count'] for s in group),
        'no_movement_count': sum(s['no_movement_count'] for s in group),
        'incorrect_count': sum(s['incorrect_count'] for s in group),
        'wheel_abort_count': sum(s['wheel_abort_count'] for s in group),
        'l_count': sum(s['l_count'] for s in group),
        'r_count': sum(s['r_count'] for s in group),
        'lick_count': sum(s['lick_count'] for s in group),
        'threshold_start_deg': first['threshold_start_deg'],
        'threshold_end_deg': last['threshold_end_deg'],
        'threshold_final_deg': last['threshold_final_deg'],
        'iti_end_s': last['iti_end_s'],
        'gain_mult_end': last['gain_mult_end'],
        'direction_ratio_end': last['direction_ratio_end'],
        'simple_gates_met': last['simple_gates_met'],
        'session_csv_path': '; '.join(s['session_csv_path'] for s in group),
        'session_struct_path': None,   # filled in by the caller
    }


def _stage_sort_key(protocol_name):
    """ Ascending stage-number order (stage2_... -> 2); anything not matching the stageN_
    convention sorts alphabetically after every numbered stage. """
    match = re.match(r'stage(\d+)', protocol_name)
    if match:
        return (0, int(match.group(1)), protocol_name)
    return (1, 0, protocol_name)


_TEST_PROTOCOL_LABELS = {
    'full_protocol_lookback_test': 'Test',   # the Tests-project protocol currently scanned -- see
                                              # summarize_session()'s docstring. Add further
                                              # entries here if more Tests-project protocols ever
                                              # get a real PROTOCOL_CONFIG entry.
}


def _stage_label(protocol_name):
    match = re.match(r'stage(\d+)', protocol_name)
    if match:
        return 'Stage {0}'.format(match.group(1))
    return _TEST_PROTOCOL_LABELS.get(protocol_name, protocol_name)


# --- workbook update-in-place --------------------------------------------------------------------

# ('internal key', 'column header label'). This is the CANONICAL order -- every run rewrites the
# entire table (header row included) to match it exactly, so a reorder here actually takes effect
# on an existing sheet too, not just a brand-new one. No 'protocol' column: which stage a row
# belongs to is already unambiguous from the section header it sits under (see
# _write_section_header()), so a per-row repeat of the same text would just be redundant.
COLUMNS = [
    ('date', 'Date'),
    ('time_of_day', 'Time (HH:MM)'),
    ('weight_g', 'Weight (g)'),                    # MANUAL -- never written for an existing row
    ('weight_pct', 'Weight %'),                    # formula, refreshed every run
    ('weight_after_task_g', 'Weight after task (g)'),   # MANUAL, same treatment as weight_g
    ('weight_after_task_pct', 'Weight after task %'),   # formula, same treatment as weight_pct --
                                                         # references weight_after_task_g instead
    ('hand_watered', 'Water given by hand after task'), # MANUAL, TRUE/FALSE
    ('trial_count', 'Trials'),
    ('session_duration_s', 'Duration (s)'),
    ('reward_ul', 'Reward Amount (uL)'),            # auto, from the REWARD_UL VAL
    ('consumed_volume_ul', 'Total Volume (uL)'),    # auto, only counts licked/consumed rewards
    ('aborts', 'Aborts'),
    ('withheld_count', 'Withheld'),
    ('no_movement_count', 'No-movement'),
    ('incorrect_count', 'Incorrect'),           # full_protocol_lookback_test only -- a genuine
                                                 # wrong-choice outcome, distinct from Withheld
    ('wheel_abort_count', 'Wheel Aborts'),      # full_protocol_lookback_test only -- mid-cue wheel
                                                 # movement aborts, distinct from Aborts (which is
                                                 # Stage 1/2's QUIESCENCE_BREAKS count)
    ('l_count', 'L turns'),
    ('r_count', 'R turns'),
    ('lick_count', 'Licks'),
    ('threshold_start_deg', 'Threshold start (deg)'),
    ('threshold_end_deg', 'Threshold end (deg)'),
    ('threshold_pct_final', 'Threshold end (% final)'),
    ('iti_end_s', 'ITI (s)'),
    ('gain_mult_end', 'Gain mult (Stage 1)'),
    ('direction_ratio_end', 'Direction ratio (Stage 2)'),
    ('simple_gates_met', 'Simple gates met (Stage 2)'),
    ('num_sessions', 'Num. Sessions'),
    ('notes', 'Notes'),                       # MANUAL -- never written for an existing row
    ('session_started', 'Session started'),   # hidden key -- `; `-joined member timestamps
    ('session_csv_path', 'Session data (local path)'),
    ('session_struct_path', 'Session struct (local path)'),
]
_MANUAL_COLUMNS = {'weight_g', 'notes', 'weight_after_task_g', 'hand_watered'}

_HEADER_FIELD_LABELS = ['DOB', 'Sex', 'Strain', 'Baseline weight (g)']
_BASELINE_WEIGHT_ROW = 2 + _HEADER_FIELD_LABELS.index('Baseline weight (g)')   # = 5
_BASELINE_WEIGHT_COL_LETTER = 'B'

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=12)
_SECTION_FONT = Font(bold=True, size=11)

# One distinct fill per stage number (1-indexed into this list, wrapping if there are ever more
# stages than colors) -- blue/green/yellow/orange/purple, all light enough for black text to stay
# readable. Anything that doesn't match the stageN_ naming convention falls back to plain gray.
_STAGE_FILL_COLORS = ['DDEBF7', 'E2F0D9', 'FFF2CC', 'FCE4D6', 'E4DFEC']
_DEFAULT_FILL_COLOR = 'F2F2F2'


def _stage_fill_color(protocol_name):
    match = re.match(r'stage(\d+)', protocol_name)
    if not match:
        return _DEFAULT_FILL_COLOR
    stage_num = int(match.group(1))
    return _STAGE_FILL_COLORS[(stage_num - 1) % len(_STAGE_FILL_COLORS)]


def _sheet_name_for(subject):
    return ''.join(c for c in subject if c not in r'[]:*?/\\')[:31] or 'sheet'


def _get_or_create_subject_sheet(wb, subject):
    """ Returns (ws, table_start_row). If the sheet already exists, its DOB/Sex/Strain/Baseline
    weight header cells are left completely untouched (may already be hand-filled). """
    sheet_name = _sheet_name_for(subject)
    table_start_row = 2 + len(_HEADER_FIELD_LABELS) + 1
    if sheet_name in wb.sheetnames:
        return wb[sheet_name], table_start_row

    ws = wb.create_sheet(sheet_name)
    ws['A1'] = 'Animal: {0}'.format(subject)
    ws['A1'].font = _TITLE_FONT
    for i, label in enumerate(_HEADER_FIELD_LABELS):
        ws.cell(row=2 + i, column=1, value=label).font = _HEADER_FONT
        # value cell (column B) deliberately left blank for hand-entry
    return ws, table_start_row


def _canonical_col_map():
    """ Fixed column positions matching COLUMNS' own order, 1-indexed. Used for ALL writing --
    the table gets fully rewritten in this order every run (see _rewrite_subject_sheet()). """
    return {key: idx for idx, (key, _label) in enumerate(COLUMNS, start=1)}


def _scan_existing_column_labels(ws, table_start_row):
    """ Returns {key: col_idx} for whatever COLUMNS-known labels are CURRENTLY in the header row,
    at whatever position they happen to be -- used ONLY to locate old data (hand-typed manual
    cells, the hidden session-key column) before the table gets cleared and rewritten in canonical
    order. Any label not in COLUMNS (an old, now-retired column, e.g. a since-removed 'Protocol /
    Stage' or legacy 'Rewards' column from an earlier schema) is simply not reported -- its data
    doesn't survive the rewrite, which is intentional: those columns were already fully
    superseded, nothing of value is lost. """
    label_to_key = {label: key for key, label in COLUMNS}
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=table_start_row, column=col_idx).value
        if value in label_to_key:
            col_map[label_to_key[value]] = col_idx
    return col_map


def _read_existing_manual_cells(ws, table_start_row, col_map):
    """ Returns {frozenset(member session_started keys): {'weight_g':..., 'notes':...}} for every
    existing DATA row (section-header rows have no session_started value, so they're skipped
    automatically). Read BEFORE the data area is cleared, so hand-typed Weight (g)/Notes values
    can be carried forward onto whichever row now represents that same underlying session(s). """
    existing = {}
    key_col = col_map.get('session_started')
    if key_col is None:
        return existing
    weight_col = col_map.get('weight_g')
    notes_col = col_map.get('notes')
    weight_after_col = col_map.get('weight_after_task_g')
    hand_watered_col = col_map.get('hand_watered')
    for row in range(table_start_row + 1, ws.max_row + 1):
        key_val = ws.cell(row=row, column=key_col).value
        if not key_val:
            continue
        member_keys = frozenset(k.strip() for k in str(key_val).split(';') if k.strip())
        if not member_keys:
            continue
        existing[member_keys] = {
            'weight_g': ws.cell(row=row, column=weight_col).value if weight_col else None,
            'notes': ws.cell(row=row, column=notes_col).value if notes_col else None,
            'weight_after_task_g': (ws.cell(row=row, column=weight_after_col).value
                                     if weight_after_col else None),
            'hand_watered': (ws.cell(row=row, column=hand_watered_col).value
                              if hand_watered_col else None),
        }
    return existing


def _carry_forward_manual_cells(member_key_set, existing):
    """ Exact match (this exact group already had a row) carries its manual cells straight over.
    Otherwise, if some existing row's member-set is a strict SUBSET of this (bigger) group's
    member-set, that row represents a session that used to stand alone (or was a smaller group)
    and has now been absorbed into a bigger one after a restart appeared -- its manual cells still
    carry forward, since it's the same underlying training bout, just now more complete. Returns
    (manual_cells_dict, matched_existing_row: bool). """
    if member_key_set in existing:
        return existing[member_key_set], True
    for old_keys, old_manual in existing.items():
        if old_keys < member_key_set:   # strict subset
            return old_manual, True
    return {'weight_g': None, 'notes': None, 'weight_after_task_g': None,
            'hand_watered': None}, False


def _write_row(ws, row, col_map, row_data):
    row_data = dict(row_data)
    if row_data.get('threshold_end_deg') is not None and row_data.get('threshold_final_deg'):
        row_data['threshold_pct_final'] = (row_data['threshold_end_deg']
                                            / row_data['threshold_final_deg'])
    else:
        row_data['threshold_pct_final'] = None

    for key, _ in COLUMNS:
        if key in _MANUAL_COLUMNS or key in ('weight_pct', 'weight_after_task_pct'):
            continue   # manual columns: written from whatever was carried forward, below.
        col_idx = col_map[key]
        value = row_data.get(key)
        cell = ws.cell(row=row, column=col_idx, value=value)
        if key == 'threshold_pct_final' and value is not None:
            cell.number_format = '0%'

    for key in _MANUAL_COLUMNS:
        ws.cell(row=row, column=col_map[key], value=row_data.get(key))

    # weight_pct/weight_after_task_pct are formulas (not hand-typed values), safe to refresh every
    # run regardless of whether the row is new or existing -- they only ever READ weight_g/
    # weight_after_task_g + baseline, never write them.
    for weight_key, pct_key in (('weight_g', 'weight_pct'),
                                 ('weight_after_task_g', 'weight_after_task_pct')):
        weight_col_letter = get_column_letter(col_map[weight_key])
        pct_cell = ws.cell(row=row, column=col_map[pct_key])
        pct_cell.value = '=IFERROR({0}{1}/${2}${3},"")'.format(
            weight_col_letter, row, _BASELINE_WEIGHT_COL_LETTER, _BASELINE_WEIGHT_ROW)
        pct_cell.number_format = '0.0%'


def _write_section_header(ws, row, title, num_cols, fill_color):
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    fill = PatternFill('solid', fgColor=fill_color)
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def _rewrite_subject_sheet(wb, subject, sessions):
    """ Regenerates this subject's entire data area (below the header block) from a fresh scan
    every run -- see the module docstring for why a simple incremental append isn't enough once
    merging + per-stage sectioning are both in play. Every OTHER subject's sheet in the workbook
    is never touched. """
    ws, table_start_row = _get_or_create_subject_sheet(wb, subject)
    old_col_map = _scan_existing_column_labels(ws, table_start_row)
    existing = _read_existing_manual_cells(ws, table_start_row, old_col_map)

    groups = group_sessions(sessions)

    row_dicts = []
    new_count = updated_count = 0
    for group in groups:
        row = combine_group(group)
        member_key_set = frozenset(row['session_started'].split('; '))
        manual, matched = _carry_forward_manual_cells(member_key_set, existing)
        row['weight_g'] = manual.get('weight_g')
        row['notes'] = manual.get('notes')
        row['weight_after_task_g'] = manual.get('weight_after_task_g')
        row['hand_watered'] = manual.get('hand_watered')
        if matched:
            updated_count += 1
        else:
            new_count += 1

        if len(group) > 1:
            member_csv_paths = [s['session_csv_path'] for s in group]
            out_base = os.path.splitext(group[0]['session_csv_path'])[0]
            mat_path, _json_path = session_struct_export.merge_session_structs(
                member_csv_paths, out_base)
            row['session_struct_path'] = mat_path
        else:
            row['session_struct_path'] = group[0]['session_struct_path']

        row_dicts.append(row)

    buckets = {}
    for row in row_dicts:
        buckets.setdefault(row['protocol'], []).append(row)
    for rows in buckets.values():
        rows.sort(key=lambda r: r['date'])
    ordered_protocols = sorted(buckets.keys(), key=_stage_sort_key)

    # Clear the ENTIRE table -- header row included, not just the data rows below it -- and
    # rewrite it fresh in canonical order. This is what actually lets a schema/column-order change
    # (like dropping 'Protocol / Stage' and moving Weight (g)/Weight % to C/D) take effect on an
    # existing sheet, not just a brand-new one; the old scan-and-append design preserved whatever
    # historical order a sheet already had forever, which is exactly what was wrong.
    if ws.max_row >= table_start_row:
        ws.delete_rows(table_start_row, ws.max_row - table_start_row + 1)

    col_map = _canonical_col_map()
    for key, label in COLUMNS:
        cell = ws.cell(row=table_start_row, column=col_map[key], value=label)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')

    row_num = table_start_row + 1
    num_cols = len(COLUMNS)
    for protocol in ordered_protocols:
        _write_section_header(ws, row_num, _stage_label(protocol), num_cols,
                               _stage_fill_color(protocol))
        row_num += 1
        for row in buckets[protocol]:
            _write_row(ws, row_num, col_map, row)
            row_num += 1

    # ws.column_dimensions is keyed by LETTER and persists independently of cell content/row
    # deletes -- if 'session_started' sat at a different letter in an earlier schema iteration,
    # that letter's hidden=True flag lingers forever unless explicitly cleared here. Reset every
    # column's hidden flag before re-hiding only the one that's actually 'session_started' now.
    for dim in ws.column_dimensions.values():
        dim.hidden = False
    for col_idx in col_map.values():
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.column_dimensions[get_column_letter(col_map['session_started'])].hidden = True
    ws.freeze_panes = ws.cell(row=table_start_row + 1, column=1).coordinate
    return new_count, updated_count


def build_workbook():
    by_subject = scan_all_sessions()
    if not by_subject:
        print("No session CSVs found under experiments/*/setups/*/sessions/ on this machine -- "
              "nothing to update.", flush=True)
        return

    if os.path.exists(_OUTPUT_PATH):
        wb = openpyxl.load_workbook(_OUTPUT_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # drop the default blank sheet -- one real sheet per subject instead

    total_new = total_updated = 0
    for subject in sorted(by_subject):
        new_count, updated_count = _rewrite_subject_sheet(wb, subject, by_subject[subject])
        total_new += new_count
        total_updated += updated_count
        print("  {0}: {1} new row(s), {2} updated row(s)".format(subject, new_count, updated_count),
              flush=True)

    wb.save(_OUTPUT_PATH)
    print("Wrote {0} -- {1} new row(s), {2} updated row(s) across {3} subject(s) found on this "
          "machine. Every other subject's sheet in the workbook was left untouched.".format(
              _OUTPUT_PATH, total_new, total_updated, len(by_subject)), flush=True)


if __name__ == '__main__':
    build_workbook()
